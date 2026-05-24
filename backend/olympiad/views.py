from datetime import timedelta
import random
from io import BytesIO

from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Center, Branch, Subject, Level, Student, Question, Result, StudentAnswer, MentalTask, BRANCH_CHOICES, DEFAULT_BRANCHES

def is_mental_subject(subject_name):
    normalized = subject_name.lower().replace("'", '').replace('‘', '').replace('’', '')
    return 'mental' in normalized


def get_exam_duration_minutes(student):
    if is_mental_subject(student.subject.name):
        return 5
    return 30


MENTAL_ARITHMETIC_SEQUENCES = [
    [22, -11, -11, 55, 44],
    [33, 66, -99, 77, 11],
    [66, 33, -77, -11, 22],
    [33, 66, -55, 55, -11],
    [11, 66, -11, 33, -33],
    [99, -33, -55, -11, 44],
    [22, 77, -44, -55, 55],
    [55, 44, -22, -66, 66],
    [22, 22, 55, -77, 11],
    [22, 55, 22, -88, -11],
    [33, 11, 55, -11, 11],
    [77, 11, -22, 33, -66],
    [55, -55, 55, 11, 33],
    [99, -22, 22, -99, 77],
    [33, -33, 22, -22, 33],
    [66, -11, -55, 33, -33],
    [77, -55, -22, 77, 22],
    [99, -22, 22, -66, 11],
    [66, 22, 11, -22, 11],
    [11, -11, 55, -55, 99],
    [66, 33, -66, -11, 55],
    [66, -11, -55, 77, 11],
    [44, 55, -44, -55, 99],
    [33, -11, 11, 55, -88],
    [33, 55, 11, -77, 66],
    [88, -33, -55, 55, 44],
    [11, -11, 88, 11, -44],
    [33, 66, -88, 33, 55],
    [22, 66, -33, -55, 44],
    [22, 77, -88, -11, 55],
    [77, -55, 11, 66, -11],
    [55, -55, 33, 11, -44],
    [66, -55, -11, 44, 55],
    [55, -55, 88, -22, -11],
    [11, -11, 88, -88, 44],
    [66, 11, 22, -55, -22],
    [44, 55, -44, -55, 99],
    [44, -33, -11, 22, -22],
    [88, -33, -55, 22, -22],
    [55, -55, 99, -88, -11],
    [55, -55, 99, -33, 11],
    [66, 33, -77, 55, 11],
    [33, 55, -11, -22, -55],
    [99, -33, 11, 22, -44],
    [11, 33, -33, 66, -22],
    [33, 11, 55, -77, 66],
    [55, 33, 11, -22, 22],
    [33, 11, 55, -33, -55],
    [99, -11, -77, -11, 33],
    [77, 22, -22, -55, 55],
    [55, -55, 44, -33, 66],
    [99, -88, -11, 99, -44],
    [55, 11, 33, -66, 66],
    [77, 22, -11, 11, -88],
    [44, -11, 55, -88, 33],
    [66, 33, -11, 11, -33],
    [22, 77, -99, 11, -11],
    [55, -55, 66, -11, 11],
    [44, -44, 11, -11, 99],
    [22, 55, 11, 11, -33],
    [66, -55, -11, 44, -33],
    [11, 33, -33, 88, -22],
    [44, 55, -66, -11, -22],
    [77, 22, -22, 11, 11],
    [66, -11, 22, 11, 11],
    [44, 55, -11, -22, 11],
    [22, 77, -33, -66, 33],
    [33, -22, -11, 44, 55],
    [11, 88, -33, 33, -99],
    [55, 33, -88, 55, -55],
    [77, 22, -66, 66, -11],
    [22, 22, 55, -33, 11],
    [99, -33, 22, 11, -66],
    [33, -33, 99, -11, -77],
    [44, -33, -11, 66, -66],
    [66, 11, -77, 88, 11],
    [88, 11, -99, 33, -33],
    [44, -11, 66, -77, 77],
    [77, 11, 11, -88, 88],
    [77, 22, -44, -55, 99],
]

def format_mental_number(value, is_first=False):
    return str(value)


def generate_mental_tasks(student, count=None):
    existing_tasks = list(MentalTask.objects.filter(student=student).order_by('task_order'))
    if existing_tasks:
        return existing_tasks

    created = []
    sequences = MENTAL_ARITHMETIC_SEQUENCES if count is None else MENTAL_ARITHMETIC_SEQUENCES[:count]

    for order, sequence in enumerate(sequences, start=1):
        flashes = [format_mental_number(value, index == 0) for index, value in enumerate(sequence)]
        correct_answer = sum(sequence)
        task = MentalTask.objects.create(
            student=student,
            task_order=order,
            flashes=flashes,
            expression=' '.join(flashes),
            correct_answer=correct_answer,
        )
        created.append(task)
    return created





from .serializers import (
    CenterSerializer,
    BranchSerializer,
    SubjectSerializer,
    LevelSerializer,
    StudentSerializer,
    QuestionAdminSerializer,
    QuestionForExamSerializer,
    ResultSerializer,
)


class IsAdminUser(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated and request.user.is_staff)


def get_user_profile(user):
    return getattr(user, 'admin_profile', None) if user and user.is_authenticated else None


def get_user_branch(user):
    return getattr(get_user_profile(user), 'branch', '')


def get_user_center(user):
    profile = get_user_profile(user)
    return getattr(profile, 'center', None)


def is_main_admin(user):
    profile = get_user_profile(user)
    return bool(user and user.is_authenticated and user.is_staff and (user.is_superuser or not profile or (not profile.center_id and not profile.branch)))


def can_manage_students(user):
    return bool(is_main_admin(user) or get_user_center(user))


def branch_key(text):
    return str(text or '').strip().lower().replace("'", '').replace('‘', '').replace('’', '').replace('`', '')


def normalize_branch(value):
    clean = str(value or '').strip()
    if not clean:
        return ''

    branch_map = {branch_key(branch): branch for branch in DEFAULT_BRANCHES}
    try:
        branch_map.update({branch_key(branch.name): branch.name for branch in Branch.objects.all()})
    except Exception:
        pass
    return branch_map.get(branch_key(clean), clean)


def ensure_default_branches():
    for branch_name in DEFAULT_BRANCHES:
        Branch.objects.get_or_create(name=branch_name)


class CenterViewSet(viewsets.ModelViewSet):
    queryset = Center.objects.all()
    serializer_class = CenterSerializer
    permission_classes = [IsAdminUser]

    def get_queryset(self):
        qs = super().get_queryset()
        user_center = get_user_center(self.request.user)
        if not is_main_admin(self.request.user) and user_center:
            qs = qs.filter(id=user_center.id)
        elif not is_main_admin(self.request.user):
            qs = qs.none()
        return qs.order_by('name')

    def create(self, request, *args, **kwargs):
        if not is_main_admin(request.user):
            return Response({'detail': 'Sizga o‘quv markaz qo‘shishga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        name = str(request.data.get('name') or '').strip()
        if not name:
            return Response({'name': 'O‘quv markaz nomini kiriting.'}, status=status.HTTP_400_BAD_REQUEST)
        existing = Center.objects.filter(name__iexact=name).first()
        if existing:
            serializer = self.get_serializer(existing)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if not is_main_admin(request.user):
            return Response({'detail': 'Sizga o‘quv markazni tahrirlashga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not is_main_admin(request.user):
            return Response({'detail': 'Sizga o‘quv markazni tahrirlashga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not is_main_admin(request.user):
            return Response({'detail': 'Sizga o‘quv markazni o‘chirishga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        center = self.get_object()
        if Student.objects.filter(center=center).exists():
            return Response({'detail': 'Bu o‘quv markazda o‘quvchilar bor. Avval ularni boshqa markazga o‘tkazing.'}, status=status.HTTP_400_BAD_REQUEST)
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        data = []
        for center in self.get_queryset():
            students_qs = Student.objects.filter(center=center)
            results_qs = Result.objects.filter(student__center=center)
            results = list(results_qs.values_list('percent', flat=True))
            avg_percent = round(sum(results) / len(results), 1) if results else 0
            data.append({
                'id': center.id,
                'name': center.name,
                'students_count': students_qs.count(),
                'not_started_count': students_qs.filter(status=Student.Status.NOT_STARTED).count(),
                'in_progress_count': students_qs.filter(status=Student.Status.IN_PROGRESS).count(),
                'completed_count': students_qs.filter(status=Student.Status.COMPLETED).count(),
                'results_count': len(results),
                'average_percent': avg_percent,
            })
        return Response(data)


class BranchViewSet(viewsets.ModelViewSet):
    queryset = Branch.objects.all()
    serializer_class = BranchSerializer
    permission_classes = [IsAdminUser]

    def list(self, request, *args, **kwargs):
        ensure_default_branches()
        return super().list(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        if not is_main_admin(request.user):
            return Response({'detail': 'Sizga filial qo‘shishga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        name = str(request.data.get('name') or '').strip()
        if not name:
            return Response({'name': 'Filial nomini kiriting.'}, status=status.HTTP_400_BAD_REQUEST)
        existing = Branch.objects.filter(name__iexact=name).first()
        if existing:
            serializer = self.get_serializer(existing)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if not is_main_admin(request.user):
            return Response({'detail': 'Sizga filialni tahrirlashga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not is_main_admin(request.user):
            return Response({'detail': 'Sizga filialni tahrirlashga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not is_main_admin(request.user):
            return Response({'detail': 'Sizga filialni o‘chirishga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        branch = self.get_object()
        if Student.objects.filter(branch=branch.name).exists():
            return Response({'detail': 'Bu filialda o‘quvchilar bor. Avval ularni boshqa filialga o‘tkazing.'}, status=status.HTTP_400_BAD_REQUEST)
        return super().destroy(request, *args, **kwargs)


class SubjectViewSet(viewsets.ModelViewSet):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = [IsAdminUser]


class LevelViewSet(viewsets.ModelViewSet):
    queryset = Level.objects.select_related('subject').all()
    serializer_class = LevelSerializer
    permission_classes = [IsAdminUser]

    def get_queryset(self):
        qs = super().get_queryset()
        subject_id = self.request.query_params.get('subject')
        if subject_id:
            qs = qs.filter(subject_id=subject_id)
        return qs


class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.select_related('subject', 'level', 'center').all()
    serializer_class = StudentSerializer
    permission_classes = [IsAdminUser]

    def get_queryset(self):
        qs = super().get_queryset()
        q = self.request.query_params.get('q')
        status_value = self.request.query_params.get('status')
        subject = self.request.query_params.get('subject')
        level = self.request.query_params.get('level')
        center = self.request.query_params.get('center')
        branch = normalize_branch(self.request.query_params.get('branch'))
        user_center = get_user_center(self.request.user)
        user_branch = get_user_branch(self.request.user)

        if not is_main_admin(self.request.user) and user_center:
            qs = qs.filter(center=user_center)
        elif not is_main_admin(self.request.user) and user_branch:
            qs = qs.filter(branch=user_branch)
        elif center:
            qs = qs.filter(center_id=center)

        if branch and is_main_admin(self.request.user):
            qs = qs.filter(branch=branch)

        if q:
            qs = qs.filter(Q(first_name__icontains=q) | Q(last_name__icontains=q) | Q(code__icontains=q))
        if status_value:
            qs = qs.filter(status=status_value)
        if subject:
            qs = qs.filter(subject_id=subject)
        if level:
            qs = qs.filter(level_id=level)
        return qs.order_by('-created_at')

    def create(self, request, *args, **kwargs):
        if not can_manage_students(request.user):
            return Response({'detail': 'Sizga o‘quvchi yaratishga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        user_center = get_user_center(self.request.user)
        branch = normalize_branch(serializer.validated_data.get('branch')) or get_user_branch(self.request.user) or 'Boshqa'
        if not is_main_admin(self.request.user) and user_center:
            serializer.save(center=user_center, branch=branch)
        else:
            serializer.save(branch=branch)

    def update(self, request, *args, **kwargs):
        if not can_manage_students(request.user):
            return Response({'detail': 'Sizga o‘quvchini tahrirlashga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not can_manage_students(request.user):
            return Response({'detail': 'Sizga o‘quvchini tahrirlashga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not can_manage_students(request.user):
            return Response({'detail': 'Sizga o‘quvchini o‘chirishga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['post'], url_path='import-excel')
    def import_excel(self, request):
        if not can_manage_students(request.user):
            return Response({'detail': 'Sizga Excel orqali o‘quvchi yaratishga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'detail': 'Excel file yuborilmadi.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = load_workbook(file_obj)
            ws = wb.active
        except Exception:
            return Response({'detail': 'Excel faylni o‘qib bo‘lmadi.'}, status=status.HTTP_400_BAD_REQUEST)

        def normalize_header(value):
            text = str(value or '').strip().lower()
            replacements = {
                '‘': "'",
                '’': "'",
                '`': "'",
                'ʼ': "'",
                '№': 'no',
                '#': 'no',
            }
            for old, new_value in replacements.items():
                text = text.replace(old, new_value)
            text = ' '.join(text.split())
            return text

        headers = [normalize_header(cell.value) for cell in ws[1]]
        header_map = {name: idx for idx, name in enumerate(headers) if name}

        def cell_value(row, names):
            for name in names:
                idx = header_map.get(normalize_header(name))
                if idx is not None and idx < len(row):
                    value = row[idx]
                    return str(value).strip() if value is not None else ''
            return ''

        def split_full_name(full_name):
            parts = str(full_name or '').strip().split()
            if not parts:
                return '', ''
            if len(parts) == 1:
                return parts[0], ''
            return parts[0], ' '.join(parts[1:])

        created = []
        errors = []
        user_center = get_user_center(request.user)
        user_branch = get_user_branch(request.user)

        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                full_name = cell_value(row, [
                    'Ism familya',
                    'Ism familyasi',
                    'Ism Familyasi',
                    'F.I.Sh',
                    'FISH',
                    'FIO',
                    'Oquvchi',
                    "O'quvchi",
                    'O‘quvchi',
                ])
                first_name = cell_value(row, ['Ism', 'first_name', 'First name'])
                last_name = cell_value(row, ['Familya', 'last_name', 'Last name'])

                if full_name and not first_name and not last_name:
                    first_name, last_name = split_full_name(full_name)

                subject_name = cell_value(row, ['Fan', 'Subject'])
                level_name = cell_value(row, ['Daraja', 'Level'])
                center_name = cell_value(row, ["O'quv markaz", 'O‘quv markaz', 'Oquv markaz', "O'quv markazi", 'O‘quv markazi', 'Oquv markazi', 'Center'])
                branch_name = normalize_branch(cell_value(row, ['Filial', 'Branch']))

                if not is_main_admin(request.user) and user_center:
                    center = user_center
                    center_name = user_center.name
                    if not branch_name:
                        branch_name = user_branch
                else:
                    center = None

                if not branch_name:
                    branch_name = 'Boshqa'

                if not any([full_name, first_name, last_name, subject_name, level_name, center_name, branch_name]):
                    continue

                if not all([first_name, subject_name, level_name, center_name, branch_name]):
                    errors.append({
                        'row': row_num,
                        'error': "Majburiy ustunlar: Ism familya, Fan, Daraja, O'quv markaz, Filial.",
                    })
                    continue

                allowed_branches = {branch_key(branch.name) for branch in Branch.objects.all()} | {branch_key(branch) for branch in DEFAULT_BRANCHES}
                if branch_key(branch_name) not in allowed_branches:
                    errors.append({
                        'row': row_num,
                        'error': f"Filial noto‘g‘ri: {branch_name}",
                    })
                    continue

                subject, _ = Subject.objects.get_or_create(name=subject_name)
                level, _ = Level.objects.get_or_create(subject=subject, name=level_name, defaults={'duration_minutes': 30})
                if center is None:
                    center, _ = Center.objects.get_or_create(name=center_name)

                student = Student.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    subject=subject,
                    level=level,
                    center=center,
                    branch=branch_name,
                )
                created.append(student)

        return Response({
            'created_count': len(created),
            'errors': errors,
            'students': StudentSerializer(created, many=True).data,
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        students = self.get_queryset().select_related('subject', 'level', 'center', 'result')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Barcha oquvchilar'

        headers = ['№', 'Ism familyasi', 'Fani', 'Darajasi', "O'quv markazi", 'Filial', 'Code', 'Status', 'Natijasi']
        ws.append(headers)

        header_fill = PatternFill('solid', fgColor='1F4E79')
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        status_labels = {
            Student.Status.NOT_STARTED: 'Ishlamagan',
            Student.Status.IN_PROGRESS: 'Ishlayapti',
            Student.Status.COMPLETED: 'Ishlab bo‘ldi',
        }

        for idx, student in enumerate(students, start=1):
            try:
                result = student.result
            except ObjectDoesNotExist:
                result = None

            natija = '—'
            if result:
                natija = f'{result.correct_count}/{result.total_questions} ta / {result.percent:.1f}%'

            ws.append([
                idx,
                student.full_name,
                student.subject.name,
                student.level.name,
                student.center.name,
                student.branch,
                student.code,
                status_labels.get(student.status, student.status),
                natija,
            ])

        ws.freeze_panes = 'A2'
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 4, 42)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="barcha_oquvchilar.xlsx"'
        return response

class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.objects.select_related('subject', 'level').all()
    serializer_class = QuestionAdminSerializer
    permission_classes = [IsAdminUser]

    def get_queryset(self):
        qs = super().get_queryset()
        subject = self.request.query_params.get('subject')
        level = self.request.query_params.get('level')
        if subject:
            qs = qs.filter(subject_id=subject)
        if level:
            qs = qs.filter(level_id=level)
        return qs.order_by('subject__name', 'level__name', 'id')

    def create(self, request, *args, **kwargs):
        if not is_main_admin(request.user):
            return Response({'detail': 'Sizga test qo‘shishga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if not is_main_admin(request.user):
            return Response({'detail': 'Sizga testni tahrirlashga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not is_main_admin(request.user):
            return Response({'detail': 'Sizga testni tahrirlashga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not is_main_admin(request.user):
            return Response({'detail': 'Sizga testni o‘chirishga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


class ResultViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Result.objects.select_related('student', 'student__subject', 'student__level', 'student__center').prefetch_related('answers', 'mental_answers').all()
    serializer_class = ResultSerializer
    permission_classes = [IsAdminUser]

    def get_queryset(self):
        qs = super().get_queryset()
        center = self.request.query_params.get('center')
        branch = normalize_branch(self.request.query_params.get('branch'))
        user_center = get_user_center(self.request.user)
        user_branch = get_user_branch(self.request.user)

        if not is_main_admin(self.request.user) and user_center:
            qs = qs.filter(student__center=user_center)
        elif not is_main_admin(self.request.user) and user_branch:
            qs = qs.filter(student__branch=user_branch)
        elif center:
            qs = qs.filter(student__center_id=center)

        if branch and is_main_admin(self.request.user):
            qs = qs.filter(student__branch=branch)
        return qs.order_by('-created_at')

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Natijalar'

        headers = [
            '№', 'Ism Familya', 'Fan', 'Daraja', "O'quv markaz", 'Filial', 'Status code',
            'To‘g‘ri javoblar', 'Jami savollar', 'Foiz', 'Boshlangan vaqti',
            'Tugatgan vaqti', 'Sarflagan vaqt'
        ]
        ws.append(headers)

        header_fill = PatternFill('solid', fgColor='1F4E79')
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for idx, result in enumerate(self.get_queryset(), start=1):
            spent = str(timedelta(seconds=result.spent_seconds))
            ws.append([
                idx,
                result.student.full_name,
                result.student.subject.name,
                result.student.level.name,
                result.student.center.name,
                result.student.branch,
                result.student.code,
                result.correct_count,
                result.total_questions,
                f'{result.percent:.1f}%',
                timezone.localtime(result.started_at).strftime('%Y-%m-%d %H:%M:%S') if result.started_at else '',
                timezone.localtime(result.finished_at).strftime('%Y-%m-%d %H:%M:%S') if result.finished_at else '',
                spent,
            ])

        for column_cells in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 3, 40)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="olimpiada_natijalari.xlsx"'
        return response


    @action(detail=False, methods=['get'], url_path='mental-answers')
    def mental_answers(self, request):
        qs = self.get_queryset().filter(mental_answers__isnull=False).distinct().prefetch_related('mental_answers')
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='mental-answers-export')
    def mental_answers_export(self, request):
        qs = self.get_queryset().filter(mental_answers__isnull=False).distinct().prefetch_related('mental_answers')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Mental javoblari'

        headers = [
            '№', 'Ism Familya', 'Fan', 'Daraja', "O'quv markaz", 'Filial', 'Status code',
            'Umumiy natija', 'Foiz', 'Sarflagan vaqt', 'Misol №', 'Misol', "O'quvchi javobi",
            "To'g'ri javob", 'Holat'
        ]
        ws.append(headers)

        header_fill = PatternFill('solid', fgColor='1F4E79')
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        row_index = 1
        for result in qs:
            tasks = list(result.mental_answers.all().order_by('task_order'))
            if not tasks:
                continue
            spent = str(timedelta(seconds=result.spent_seconds))
            for task in tasks:
                ws.append([
                    row_index,
                    result.student.full_name,
                    result.student.subject.name,
                    result.student.level.name,
                    result.student.center.name,
                    result.student.branch,
                    result.student.code,
                    f'{result.correct_count}/{result.total_questions}',
                    f'{result.percent:.1f}%',
                    spent,
                    task.task_order,
                    task.expression,
                    task.student_answer if task.student_answer is not None else '',
                    task.correct_answer,
                    "To'g'ri" if task.is_correct else "Noto'g'ri",
                ])
            row_index += 1

        for column_cells in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 3, 45)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="mental_javoblari.xlsx"'
        return response


class PublicResultLookupAPIView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        code = str(request.data.get('code', '')).strip()
        if not code:
            return Response({'detail': 'Status code kiriting.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            student = Student.objects.select_related('subject', 'level', 'center').get(code=code)
        except Student.DoesNotExist:
            return Response({'detail': 'Bunday code topilmadi.'}, status=status.HTTP_404_NOT_FOUND)

        if student.status != Student.Status.COMPLETED:
            return Response({'detail': 'Bu o‘quvchi hali testni yakunlamagan.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            result = student.result
        except ObjectDoesNotExist:
            return Response({'detail': 'Bu code uchun natija topilmadi.'}, status=status.HTTP_404_NOT_FOUND)

        is_mental = is_mental_subject(student.subject.name)
        mental_answers = []
        if is_mental:
            mental_answers = [
                {
                    'id': task.id,
                    'task_order': task.task_order,
                    'expression': task.expression,
                    'correct_answer': task.correct_answer,
                    'student_answer': task.student_answer,
                    'is_correct': task.is_correct,
                }
                for task in result.mental_answers.all().order_by('task_order')
            ]

        return Response({
            'student_full_name': student.full_name,
            'student_code': student.code,
            'subject_name': student.subject.name,
            'level_name': student.level.name,
            'center_name': student.center.name,
            'branch': student.branch,
            'total_questions': result.total_questions,
            'correct_count': result.correct_count,
            'percent': result.percent,
            'spent_seconds': result.spent_seconds,
            'finished_at': result.finished_at,
            'is_mental': is_mental,
            'mental_answers': mental_answers,
        })

class ExamStartAPIView(APIView):
    permission_classes = [permissions.AllowAny]

    @transaction.atomic
    def post(self, request):
        code = str(request.data.get('code', '')).strip()
        if not code:
            return Response({'detail': 'Status code kiriting.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            student = Student.objects.select_related('subject', 'level', 'center').select_for_update().get(code=code)
        except Student.DoesNotExist:
            return Response({'detail': 'Bunday code topilmadi.'}, status=status.HTTP_404_NOT_FOUND)

        if student.status == Student.Status.COMPLETED or student.is_used:
            return Response({'detail': 'Bu code oldin ishlatilgan.'}, status=status.HTTP_400_BAD_REQUEST)

        is_mental_exam = is_mental_subject(student.subject.name)

        # Agar o‘quvchi testni boshlab, adashib chiqib ketgan bo‘lsa,
        # code qayta kiritilganda bloklamaymiz. Test yakunlanmagan bo‘lsa,
        # aynan o‘sha started_at bilan davom ettiramiz.
        if student.status == Student.Status.IN_PROGRESS:
            started_at = student.started_at or timezone.now()

            if is_mental_exam:
                mental_tasks = generate_mental_tasks(student)
                return Response({
                    'mode': 'mental',
                    'student': StudentSerializer(student).data,
                    'duration_minutes': get_exam_duration_minutes(student),
                    'started_at': started_at,
                    'mental_tasks': [
                        {
                            'id': task.id,
                            'task_order': task.task_order,
                            'flashes': task.flashes,
                            'task_display_ms': 3000,
                        }
                        for task in mental_tasks
                    ],
                })

            questions = Question.objects.filter(subject=student.subject, level=student.level).order_by('id')
            if not questions.exists():
                return Response({'detail': 'Bu fan va daraja uchun testlar hali qo‘shilmagan.'}, status=status.HTTP_400_BAD_REQUEST)

            return Response({
                'mode': 'test',
                'student': StudentSerializer(student).data,
                'duration_minutes': get_exam_duration_minutes(student),
                'started_at': started_at,
                'questions': QuestionForExamSerializer(questions, many=True).data,
            })

        now = timezone.now()
        student.status = Student.Status.IN_PROGRESS
        student.started_at = now
        student.save(update_fields=['status', 'started_at'])

        if is_mental_exam:
            mental_tasks = generate_mental_tasks(student)
            return Response({
                'mode': 'mental',
                'student': StudentSerializer(student).data,
                'duration_minutes': get_exam_duration_minutes(student),
                'started_at': now,
                'mental_tasks': [
                    {
                        'id': task.id,
                        'task_order': task.task_order,
                        'flashes': task.flashes,
                        'task_display_ms': 3000,
                    }
                    for task in mental_tasks
                ],
            })

        questions = Question.objects.filter(subject=student.subject, level=student.level).order_by('id')
        if not questions.exists():
            student.status = Student.Status.NOT_STARTED
            student.started_at = None
            student.save(update_fields=['status', 'started_at'])
            return Response({'detail': 'Bu fan va daraja uchun testlar hali qo‘shilmagan.'}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            'mode': 'test',
            'student': StudentSerializer(student).data,
            'duration_minutes': get_exam_duration_minutes(student),
            'started_at': now,
            'questions': QuestionForExamSerializer(questions, many=True).data,
        })


class ExamSubmitAPIView(APIView):
    permission_classes = [permissions.AllowAny]

    def submit_mental(self, student, answers):
        task_ids = [item.get('task_id') for item in answers if isinstance(item, dict)]
        tasks = list(MentalTask.objects.select_for_update().filter(student=student, id__in=task_ids).order_by('task_order'))

        # Frontend yuborgan barcha ko‘rsatilgan misollar tekshiriladi.
        # Javob kiritilmagan misol noto‘g‘ri deb saqlanadi.
        answer_map = {}
        for item in answers:
            if not isinstance(item, dict):
                continue
            try:
                task_id = int(item.get('task_id'))
                answer = int(str(item.get('answer', '')).strip())
            except Exception:
                continue
            answer_map[task_id] = answer

        finished_at = timezone.now()
        started_at = student.started_at or finished_at
        spent_seconds = max(0, int((finished_at - started_at).total_seconds()))

        result = Result.objects.create(
            student=student,
            total_questions=len(tasks),
            correct_count=0,
            percent=0,
            started_at=started_at,
            finished_at=finished_at,
            spent_seconds=spent_seconds,
        )

        correct_count = 0
        for task in tasks:
            student_answer = answer_map.get(task.id)
            is_correct = student_answer == task.correct_answer
            if is_correct:
                correct_count += 1
            task.result = result
            task.student_answer = student_answer
            task.is_correct = is_correct
            task.save(update_fields=['result', 'student_answer', 'is_correct'])

        percent = (correct_count / len(tasks) * 100) if tasks else 0
        result.correct_count = correct_count
        result.percent = round(percent, 2)
        result.save(update_fields=['correct_count', 'percent'])

        student.status = Student.Status.COMPLETED
        student.finished_at = finished_at
        student.is_used = True
        student.save(update_fields=['status', 'finished_at', 'is_used'])

        return Response(ResultSerializer(result).data, status=status.HTTP_201_CREATED)

    @transaction.atomic
    def post(self, request):
        code = str(request.data.get('code', '')).strip()
        answers = request.data.get('answers', [])

        if not code:
            return Response({'detail': 'Status code yuborilmadi.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            student = Student.objects.select_related('subject', 'level', 'center').select_for_update().get(code=code)
        except Student.DoesNotExist:
            return Response({'detail': 'Bunday code topilmadi.'}, status=status.HTTP_404_NOT_FOUND)

        if student.status == Student.Status.COMPLETED or student.is_used:
            return Response({'detail': 'Bu code oldin ishlatilgan.'}, status=status.HTTP_400_BAD_REQUEST)

        if student.status != Student.Status.IN_PROGRESS:
            return Response({'detail': 'Avval testni boshlash kerak.'}, status=status.HTTP_400_BAD_REQUEST)

        if is_mental_subject(student.subject.name):
            return self.submit_mental(student, answers)

        answer_map = {}
        for item in answers:
            try:
                qid = int(item.get('question_id'))
            except Exception:
                continue
            selected = str(item.get('answer', '')).strip().upper()
            if selected in ['A', 'B', 'C', 'D']:
                answer_map[qid] = selected

        questions = list(Question.objects.filter(subject=student.subject, level=student.level).order_by('id'))
        correct_count = 0
        finished_at = timezone.now()
        started_at = student.started_at or finished_at
        spent_seconds = max(0, int((finished_at - started_at).total_seconds()))

        result = Result.objects.create(
            student=student,
            total_questions=len(questions),
            correct_count=0,
            percent=0,
            started_at=started_at,
            finished_at=finished_at,
            spent_seconds=spent_seconds,
        )

        answer_objects = []
        for question in questions:
            selected = answer_map.get(question.id, '')
            is_correct = selected == question.correct_answer
            if is_correct:
                correct_count += 1
            answer_objects.append(StudentAnswer(
                result=result,
                question=question,
                selected_answer=selected,
                is_correct=is_correct,
            ))
        StudentAnswer.objects.bulk_create(answer_objects)

        percent = (correct_count / len(questions) * 100) if questions else 0
        result.correct_count = correct_count
        result.percent = round(percent, 2)
        result.save(update_fields=['correct_count', 'percent'])

        student.status = Student.Status.COMPLETED
        student.finished_at = finished_at
        student.is_used = True
        student.save(update_fields=['status', 'finished_at', 'is_used'])

        return Response(ResultSerializer(result).data, status=status.HTTP_201_CREATED)
